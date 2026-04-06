"""category_mapper のテスト"""

import sys
from pathlib import Path

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))


@pytest.fixture
def mapping_file(tmp_path, monkeypatch):
    """テスト用の mapping.xlsx を作成し、CATEGORY_NAME_FILE も空ファイルで用意する"""
    fp = tmp_path / "mapping.xlsx"
    wb = openpyxl.Workbook()

    # --- Sheet1: ブランド別マッピング ---
    ws1 = wb.active
    ws1.title = "ブランド別マッピング"
    # ヘッダー (A-L列: 12列)
    ws1.append([
        "ブランド英字", "ブランドカナ", "ブランド別名", "型番",
        "シリーズ英字", "シリーズカナ", "カテゴリID", "性別",
        "col8_unused", "キーワード", "col10_unused", "追加単語",
    ])
    # ブランド+シリーズ行 (性別あり)
    ws1.append([
        "SEIKO", "セイコー", "", "",
        "PRESAGE", "プレザージュ", "CAT001", "メンズ",
        "", "", "", "腕時計",
    ])
    # ブランド+シリーズ行 (性別なし = 空)
    ws1.append([
        "SEIKO", "セイコー", "", "",
        "PROSPEX", "プロスペックス", "CAT002", "",
        "", "", "", "ダイバーズ",
    ])
    # フォールバック行 (ブランドのみ)
    ws1.append([
        "SEIKO", "セイコー", "", "",
        "（その他）", "", "CAT099", "ユニセックス",
        "", "", "", "",
    ])
    # 型番行
    ws1.append([
        "CASIO", "カシオ", "", "GA-100,GA-110",
        "G-SHOCK", "ジーショック", "CAT010", "メンズ",
        "", "", "", "腕時計",
    ])
    # CASIO フォールバック
    ws1.append([
        "CASIO", "カシオ", "", "",
        "（その他）", "", "CAT019", "",
        "", "", "", "",
    ])

    # --- Sheet2: 汎用カテゴリ ---
    ws2 = wb.create_sheet("汎用カテゴリ")
    # ヘッダー (A-F列: 6列)
    ws2.append(["性別", "ムーブメント", "針タイプ", "カテゴリID", "col4_unused", "追加単語"])
    # メンズ+クォーツ
    ws2.append(["メンズ", "クォーツ", "3針（時、分、秒）", "GEN001", "", "腕時計"])
    # レディース+クォーツ
    ws2.append(["レディース", "クォーツ", "3針（時、分、秒）", "GEN002", "", "レディース腕時計"])
    # メンズ+自動巻き
    ws2.append(["メンズ", "自動巻き", "", "GEN003", "", "機械式時計"])
    # 追加単語なし
    ws2.append(["ユニセックス", "クォーツ", "", "GEN004", "", ""])

    wb.save(fp)

    # カテゴリ名ファイル（空）
    cat_fp = tmp_path / "category_names.xlsx"
    wb2 = openpyxl.Workbook()
    ws_cat = wb2.active
    ws_cat.append(["カテゴリID", "カテゴリ名"])
    ws_cat.append(["CAT001", "セイコープレザージュ"])
    wb2.save(cat_fp)

    # config のパスをモンキーパッチ
    import config
    monkeypatch.setattr(config, "MAPPING_FILE", fp)
    monkeypatch.setattr(config, "CATEGORY_NAME_FILE", cat_fp)

    return fp


@pytest.fixture
def mapper(mapping_file):
    from modules.category_mapper import CategoryMapper
    return CategoryMapper(mapping_file)


class TestLookup:
    """lookup メソッドの基本テスト"""

    def test_brand_series_match(self, mapper):
        cat_id, level, entry = mapper.lookup(brand_en="SEIKO", series_en="PRESAGE")
        assert cat_id == "CAT001"
        assert level == "brand+series"
        assert entry["gender"] == "メンズ"

    def test_brand_only_fallback(self, mapper):
        cat_id, level, entry = mapper.lookup(brand_en="SEIKO", series_en="UNKNOWN_SERIES")
        assert cat_id == "CAT099"
        assert level == "brand_only"

    def test_model_number_match(self, mapper):
        cat_id, level, entry = mapper.lookup(brand_en="CASIO", model_number="GA-100")
        assert cat_id == "CAT010"
        assert level == "model_number"

    def test_generic_fallback(self, mapper):
        """ブランドなし → 汎用カテゴリ"""
        cat_id, level, entry = mapper.lookup(
            brand_en="", gender="メンズ", movement_type="クォーツ", hand_count="3針",
        )
        assert cat_id == "GEN001"
        assert level == "generic"

    def test_unknown(self, mapper):
        cat_id, level, entry = mapper.lookup(brand_en="")
        assert cat_id == ""
        assert level == "unknown"
        assert entry is None


class TestGetAdditionalWord:
    """get_additional_word のテスト（ブランド別マッピングのみ）"""

    def test_brand_model_match(self, mapper):
        word = mapper.get_additional_word("CASIO", "GA-100")
        assert word == "腕時計"

    def test_brand_only_match(self, mapper):
        word = mapper.get_additional_word("SEIKO", "UNKNOWN_MODEL")
        assert word == "腕時計"  # SEIKO の最初のエントリの追加単語

    def test_no_match(self, mapper):
        word = mapper.get_additional_word("UNKNOWN_BRAND", "UNKNOWN_MODEL")
        assert word == ""


class TestGenericAdditionalWord:
    """仕様: 汎用カテゴリの追加単語が正しく返される"""

    def test_generic_entry_has_additional_word(self, mapper):
        """汎用カテゴリのエントリに追加単語が含まれる"""
        cat_id, level, entry = mapper.lookup(
            brand_en="", gender="メンズ", movement_type="クォーツ", hand_count="3針",
        )
        assert level == "generic"
        assert entry["additional_word"] == "腕時計"

    def test_generic_entry_additional_word_ladies(self, mapper):
        cat_id, level, entry = mapper.lookup(
            brand_en="", gender="レディース", movement_type="クォーツ", hand_count="3針",
        )
        assert level == "generic"
        assert entry["additional_word"] == "レディース腕時計"

    def test_generic_entry_additional_word_empty(self, mapper):
        """追加単語が空の汎用カテゴリ"""
        cat_id, level, entry = mapper.lookup(
            brand_en="", gender="ユニセックス", movement_type="クォーツ",
        )
        assert level == "generic"
        assert entry["additional_word"] == ""

    def test_generic_auto_mechanical(self, mapper):
        cat_id, level, entry = mapper.lookup(
            brand_en="", gender="メンズ", movement_type="自動巻き",
        )
        assert level == "generic"
        assert entry["additional_word"] == "機械式時計"


class TestBrandGender:
    """仕様: ブランドあり時、マッピングの性別が空なら空文字が返る"""

    def test_brand_series_gender_present(self, mapper):
        _, _, entry = mapper.lookup(brand_en="SEIKO", series_en="PRESAGE")
        assert entry["gender"] == "メンズ"

    def test_brand_series_gender_empty(self, mapper):
        """マッピングの性別欄が空 → 空文字列"""
        _, _, entry = mapper.lookup(brand_en="SEIKO", series_en="PROSPEX")
        assert entry["gender"] == ""

    def test_brand_only_gender(self, mapper):
        _, _, entry = mapper.lookup(brand_en="CASIO", series_en="UNKNOWN")
        # CASIO フォールバックの性別は空
        assert entry["gender"] == ""
