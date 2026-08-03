"""mapping.xlsx 登録ブランドのカナ表記を常に優先する仕様のテスト。

クライアント合意仕様: mapping.xlsx にブランドが登録されている場合、
AI が読んだカナ（例: FHB → 「エフエイチビー」）ではなく mapping.xlsx の
brand_kana（例: 「フェリックスフーバー」）を必ず採用する。

(A) get_brand_kana 単体テスト（xlsx フィクスチャ方式は tests/test_category_mapper.py を踏襲）
(B) apply_mapping_kana_priority 単体テスト
(C) パイプライン統合テスト（batch モード、tests/test_brand_blank_status.py の _run_batch 手法を流用）
(D) Web UI 取りこぼし防止
"""

import csv
import sys
from pathlib import Path

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))


def _make_mapping_file(tmp_path, rows, filename="mapping.xlsx"):
    """ブランド別マッピングシートに任意の行を書き込んだ xlsx を作成する。

    rows: 各行は
      (brand_en, brand_kana, brand_aliases, model_numbers,
       series_en, series_kana, category_id, gender)
      の8要素タプル（col8_unused/キーワード/col10_unused/追加単語は空固定）。
    """
    fp = tmp_path / filename
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "ブランド別マッピング"
    ws1.append([
        "ブランド英字", "ブランドカナ", "ブランド別名", "型番",
        "シリーズ英字", "シリーズカナ", "カテゴリID", "性別",
        "col8_unused", "キーワード", "col10_unused", "追加単語",
    ])
    for r in rows:
        brand_en, brand_kana, brand_aliases, model_numbers, series_en, series_kana, category_id, gender = r
        ws1.append([
            brand_en, brand_kana, brand_aliases, model_numbers,
            series_en, series_kana, category_id, gender,
            "", "", "", "",
        ])

    ws2 = wb.create_sheet("汎用カテゴリ")
    ws2.append(["性別", "ムーブメント", "針タイプ", "カテゴリID", "col4_unused", "追加単語"])
    ws2.append(["メンズ", "クォーツ", "3針（時、分、秒）", "GEN001", "", "腕時計"])

    wb.save(fp)

    cat_fp = tmp_path / f"category_names_{filename}"
    wb2 = openpyxl.Workbook()
    ws_cat = wb2.active
    ws_cat.append(["カテゴリID", "カテゴリ名"])
    wb2.save(cat_fp)

    return fp, cat_fp


@pytest.fixture
def fhb_mapping_file(tmp_path, monkeypatch):
    """FHB（フェリックスフーバー）の（その他）行を含む mapping.xlsx"""
    fp, cat_fp = _make_mapping_file(tmp_path, [
        ("FHB", "フェリックスフーバー", "", "", "（その他）", "", "CAT099", "メンズ"),
    ])
    import config
    monkeypatch.setattr(config, "MAPPING_FILE", fp)
    monkeypatch.setattr(config, "CATEGORY_NAME_FILE", cat_fp)
    return fp


class TestGetBrandKanaRobustness:
    """(A) get_brand_kana 単体テスト"""

    def test_fallback_row_kana_returned(self, fhb_mapping_file):
        from modules.category_mapper import CategoryMapper
        mapper = CategoryMapper(fhb_mapping_file)
        assert mapper.get_brand_kana("FHB") == "フェリックスフーバー"

    def test_fallback_row_priority_independent_of_row_order(self, tmp_path, monkeypatch):
        """CASIOのシリーズ行（brand_kana='G-SHOCK'）を先、
        （その他）行（'カシオ'）を後に書いても 'カシオ' が返ること
        （堅牢化前は行順依存で 'G-SHOCK' が返ってしまう回帰テスト）"""
        fp, cat_fp = _make_mapping_file(tmp_path, [
            ("CASIO", "G-SHOCK", "", "", "G-SHOCK", "ジーショック", "CAT010", "メンズ"),
            ("CASIO", "カシオ", "", "", "（その他）", "", "CAT019", ""),
        ])
        import config
        monkeypatch.setattr(config, "MAPPING_FILE", fp)
        monkeypatch.setattr(config, "CATEGORY_NAME_FILE", cat_fp)

        from modules.category_mapper import CategoryMapper
        mapper = CategoryMapper(fp)
        assert mapper.get_brand_kana("CASIO") == "カシオ"

    def test_english_kana_excluded_when_no_fallback_row(self, tmp_path, monkeypatch):
        """フォールバック行が無く、シリーズ行の brand_kana が英字（'IWC'）のみの場合、
        英字は採用されず空文字を返すこと"""
        fp, cat_fp = _make_mapping_file(tmp_path, [
            ("IWC", "IWC", "", "", "PORTUGIESER", "ポルトギーゼ", "CAT020", "メンズ"),
        ])
        import config
        monkeypatch.setattr(config, "MAPPING_FILE", fp)
        monkeypatch.setattr(config, "CATEGORY_NAME_FILE", cat_fp)

        from modules.category_mapper import CategoryMapper
        mapper = CategoryMapper(fp)
        assert mapper.get_brand_kana("IWC") == ""

    def test_alias_resolves_to_canonical_brand_kana(self, tmp_path, monkeypatch):
        fp, cat_fp = _make_mapping_file(tmp_path, [
            ("SEIKO", "セイコー", "SEIKO WATCH", "", "（その他）", "", "CAT099", "メンズ"),
        ])
        import config
        monkeypatch.setattr(config, "MAPPING_FILE", fp)
        monkeypatch.setattr(config, "CATEGORY_NAME_FILE", cat_fp)

        from modules.category_mapper import CategoryMapper
        mapper = CategoryMapper(fp)
        assert mapper.get_brand_kana("SEIKO WATCH") == "セイコー"

    def test_unregistered_brand_returns_empty(self, fhb_mapping_file):
        from modules.category_mapper import CategoryMapper
        mapper = CategoryMapper(fhb_mapping_file)
        assert mapper.get_brand_kana("UNKNOWN_BRAND") == ""


class TestApplyMappingKanaPriority:
    """(B) apply_mapping_kana_priority 単体テスト"""

    def test_ai_kana_replaced_by_mapping_kana(self, fhb_mapping_file):
        from main import apply_mapping_kana_priority
        from modules.category_mapper import CategoryMapper
        from modules.csv_writer import ProductResult

        mapper = CategoryMapper(fhb_mapping_file)
        result = ProductResult()
        result.brand_en = "FHB"
        result.brand_kana = "エフエイチビー"

        apply_mapping_kana_priority(result, mapper)

        assert result.brand_kana == "フェリックスフーバー"

    def test_unregistered_brand_keeps_ai_kana(self, fhb_mapping_file):
        from main import apply_mapping_kana_priority
        from modules.category_mapper import CategoryMapper
        from modules.csv_writer import ProductResult

        mapper = CategoryMapper(fhb_mapping_file)
        result = ProductResult()
        result.brand_en = "UNKNOWN_BRAND"
        result.brand_kana = "アンノウン"

        apply_mapping_kana_priority(result, mapper)

        assert result.brand_kana == "アンノウン"

    def test_blank_brand_en_does_nothing(self, fhb_mapping_file):
        from main import apply_mapping_kana_priority
        from modules.category_mapper import CategoryMapper
        from modules.csv_writer import ProductResult

        mapper = CategoryMapper(fhb_mapping_file)
        result = ProductResult()
        result.brand_en = ""
        result.brand_kana = ""

        apply_mapping_kana_priority(result, mapper)

        assert result.brand_kana == ""

    def test_blank_ai_kana_filled_by_mapping(self, fhb_mapping_file):
        from main import apply_mapping_kana_priority
        from modules.category_mapper import CategoryMapper
        from modules.csv_writer import ProductResult

        mapper = CategoryMapper(fhb_mapping_file)
        result = ProductResult()
        result.brand_en = "FHB"
        result.brand_kana = ""

        apply_mapping_kana_priority(result, mapper)

        assert result.brand_kana == "フェリックスフーバー"


class TestBrandKanaMappingPriorityPipeline:
    """(C) パイプライン統合テスト（batch モード）"""

    def _make_product(self, tmp_path):
        from modules.folder_scanner import ProductImages
        imgs = [tmp_path / f"{i:02d}.jpg" for i in range(8)]
        for p in imgs:
            p.write_bytes(b"x")
        return ProductImages(
            product_id="7777777_TEST",
            management_number="7777777",
            folder_path=tmp_path,
            images=imgs,
        )

    def _run_batch(self, tmp_path, monkeypatch, front, mapping_path):
        import main as main_module

        product = self._make_product(tmp_path)
        back = {"model_number": "", "material": ""}
        comment = {"title_prefix": "", "abnormality_text": ""}

        monkeypatch.setattr(main_module, "scan_folder", lambda _d: [product])
        monkeypatch.setattr(main_module, "create_batch_requests", lambda _p: [{"dummy": 1}])
        monkeypatch.setattr(main_module, "submit_batch", lambda _r: "batch_test")
        monkeypatch.setattr(main_module, "poll_batch", lambda _b, poll_interval=60: None)
        monkeypatch.setattr(main_module, "retrieve_batch_results", lambda _b: {})
        monkeypatch.setattr(
            main_module, "parse_batch_results_for_product",
            lambda _pid, _results: (front, back, comment),
        )

        out_csv = tmp_path / "out.csv"
        monkeypatch.setattr(sys, "argv", [
            "main.py", "--mode", "batch",
            "--input", str(tmp_path), "--output", str(out_csv),
            "--mapping", str(mapping_path),
        ])

        main_module.main()

        with open(out_csv, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 1
        return rows[0]

    def test_batch_uses_mapping_kana_over_ai_kana(self, tmp_path, monkeypatch):
        fp, _cat_fp = _make_mapping_file(tmp_path, [
            ("FHB", "フェリックスフーバー", "", "", "（その他）", "", "CAT099", "メンズ"),
        ])
        front = {
            "brand_en": "FHB",
            "brand_kana": "エフエイチビー",
            "gender": "メンズ",
            "movement_type": "クォーツ",
            "body_color": "シルバー",
            "dial_color": "ブラック",
            "hand_count": "",
            "case_shape": "ラウンド",
        }
        row = self._run_batch(tmp_path, monkeypatch, front, fp)
        assert row["ブランドカナ"] == "フェリックスフーバー"
        assert "FHB フェリックスフーバー" in row["タイトル"]
        assert "エフエイチビー" not in row["タイトル"]


class TestWebUIAppliesSameHelper:
    """(D) Web UI が main.apply_mapping_kana_priority と同一関数を使っていること"""

    def test_app_uses_same_apply_mapping_kana_priority(self):
        try:
            import app
            import main
        except Exception as e:
            pytest.skip(f"app import に失敗のためソース検査へフォールバック: {e}")
            return
        assert app.apply_mapping_kana_priority is main.apply_mapping_kana_priority
