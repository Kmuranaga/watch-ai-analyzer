"""csv_writer モジュールのテスト"""

import csv
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from modules.csv_writer import ProductResult, COLUMNS, write_csv, write_excel


class TestProductResult:
    """ProductResult データクラスのテスト"""

    def test_default_values(self):
        r = ProductResult()
        assert r.management_number == ""
        assert r.status == "正常"

    def test_to_row_length(self):
        """to_row の列数が COLUMNS と一致"""
        r = ProductResult()
        assert len(r.to_row()) == len(COLUMNS)

    def test_to_row_values(self):
        r = ProductResult(
            management_number="123",
            brand_en="SEIKO",
            title="テストタイトル",
            status="正常",
        )
        row = r.to_row()
        assert row[0] == "123"  # 管理番号
        assert row[3] == "テストタイトル"  # タイトル
        assert row[4] == "SEIKO"  # ブランド英字
        assert row[-1] == "正常"  # ステータス

    def test_to_row_order_matches_columns(self):
        """to_row の順序が COLUMNS 定義と対応している"""
        r = ProductResult(
            management_number="MN",
            category_id="CID",
            category_name="CNAME",
            title="TITLE",
            brand_en="BRAND",
        )
        row = r.to_row()
        # COLUMNS[0] = "管理番号" → row[0] = management_number
        assert row[0] == "MN"
        # COLUMNS[1] = "カテゴリ番号" → row[1] = category_id
        assert row[1] == "CID"
        # COLUMNS[2] = "カテゴリ名" → row[2] = category_name
        assert row[2] == "CNAME"
        # COLUMNS[3] = "タイトル" → row[3] = title
        assert row[3] == "TITLE"


class TestWriteCsv:
    """write_csv のテスト"""

    def test_creates_file(self, tmp_path):
        out = tmp_path / "result.csv"
        write_csv([], out)
        assert out.exists()

    def test_header_row(self, tmp_path):
        out = tmp_path / "result.csv"
        write_csv([], out)
        with open(out, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert header == COLUMNS

    def test_data_rows(self, tmp_path):
        out = tmp_path / "result.csv"
        results = [
            ProductResult(management_number="001", brand_en="SEIKO"),
            ProductResult(management_number="002", brand_en="OMEGA"),
        ]
        write_csv(results, out)
        with open(out, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 3  # header + 2 data rows
        assert rows[1][0] == "001"
        assert rows[2][0] == "002"

    def test_creates_parent_dirs(self, tmp_path):
        out = tmp_path / "sub" / "dir" / "result.csv"
        write_csv([], out)
        assert out.exists()

    def test_utf8_bom_encoding(self, tmp_path):
        """BOM付きUTF-8であること"""
        out = tmp_path / "result.csv"
        write_csv([], out)
        with open(out, "rb") as f:
            bom = f.read(3)
        assert bom == b"\xef\xbb\xbf"


class TestWriteExcel:
    """write_excel のテスト"""

    def test_creates_file(self, tmp_path):
        out = tmp_path / "result.xlsx"
        write_excel([], out)
        assert out.exists()

    def test_header_in_excel(self, tmp_path):
        import openpyxl

        out = tmp_path / "result.xlsx"
        write_excel([], out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        header = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS) + 1)]
        assert header == COLUMNS
        wb.close()

    def test_data_in_excel(self, tmp_path):
        import openpyxl

        out = tmp_path / "result.xlsx"
        results = [ProductResult(management_number="001", brand_en="SEIKO", status="正常")]
        write_excel(results, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "001"
        assert ws.cell(row=2, column=5).value == "SEIKO"
        wb.close()

    def test_error_row_colored(self, tmp_path):
        """エラー行に色が付く"""
        import openpyxl

        out = tmp_path / "result.xlsx"
        results = [ProductResult(status="処理エラー: テスト")]
        write_excel(results, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00FFC7CE"  # エラー色（赤系）
        wb.close()

    def test_warning_row_colored(self, tmp_path):
        """警告行に色が付く"""
        import openpyxl

        out = tmp_path / "result.xlsx"
        results = [ProductResult(status="カテゴリ未確定")]
        write_excel(results, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00FFEB9C"  # 警告色（黄系）
        wb.close()

    def test_freeze_panes(self, tmp_path):
        import openpyxl

        out = tmp_path / "result.xlsx"
        write_excel([], out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.freeze_panes == "A2"
        wb.close()
