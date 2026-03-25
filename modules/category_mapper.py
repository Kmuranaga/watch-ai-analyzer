"""
カテゴリマッピングモジュール
mapping.xlsxを参照し、ブランド+シリーズからカテゴリ番号を決定する

6段階フォールバックロジック:
  優先度0: ブランド+型番 完全一致（最優先）
  優先度1: ブランド+シリーズ 完全一致
  優先度1.5: シリーズ前方一致（例: "G-SHOCK FROGMAN" → "G-SHOCK"）
  優先度2: ブランドのみ一致 →「（その他）」カテゴリ
  優先度3: 汎用カテゴリ（性別+ムーブメント+針数）
  優先度4: 空白（手動入力）
"""

import logging
import re
from pathlib import Path

import openpyxl

from config import MAPPING_FILE, CATEGORY_NAME_FILE

logger = logging.getLogger(__name__)


class CategoryMapper:
    """カテゴリマッピングエンジン"""

    def __init__(self, mapping_file: Path | None = None):
        self.mapping_file = mapping_file or MAPPING_FILE
        self.brand_series_map: dict[tuple[str, str], dict] = {}  # (brand, series) -> row
        self.brand_fallback_map: dict[str, dict] = {}  # brand -> fallback row
        self.keyword_map: dict[str, tuple[str, str]] = {}  # keyword -> (brand, series)
        self.model_number_map: dict[tuple[str, str], dict] = {}  # (brand, model_number) -> row
        self.generic_categories: list[dict] = []  # 汎用カテゴリ
        self.category_name_map: dict[str, str] = {}  # カテゴリ番号 -> カテゴリ名
        self._load()
        self._load_category_names()

    def _load(self):
        """mapping.xlsx を読み込む"""
        if not self.mapping_file.exists():
            raise FileNotFoundError(f"マッピングファイルが見つかりません: {self.mapping_file}")

        wb = openpyxl.load_workbook(self.mapping_file, read_only=True)

        # === Sheet1: ブランド別マッピング ===
        ws1 = wb["ブランド別マッピング"]
        header_row = True
        for row in ws1.iter_rows(min_row=1, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]

            # ヘッダー行スキップ
            if header_row:
                header_row = False
                continue

            # 空行スキップ
            if not vals or all(v == "" for v in vals):
                continue

            # セクションヘッダー行スキップ（【...】で始まる行）
            if vals[0].startswith("【"):
                continue

            brand_en = vals[0].upper() if vals[0] else ""
            brand_kana = vals[1] if len(vals) > 1 else ""
            model_numbers = vals[2] if len(vals) > 2 else ""  # 型番（カンマ区切り複数可）
            series_en = vals[3].upper() if len(vals) > 3 and vals[3] else ""
            series_kana = vals[4] if len(vals) > 4 else ""
            category_id = vals[5] if len(vals) > 5 else ""
            gender = vals[6] if len(vals) > 6 else ""
            keywords = vals[8] if len(vals) > 8 else ""
            additional_word = vals[10] if len(vals) > 10 else ""  # 追加単語

            if not brand_en:
                continue

            entry = {
                "brand_en": brand_en,
                "brand_kana": brand_kana,
                "series_en": series_en,
                "series_kana": series_kana,
                "category_id": category_id,
                "gender": gender,
                "additional_word": additional_word,
            }

            # フォールバック行（「（その他）」）
            if series_en in ("（その他）", "(その他)", "（その他）".upper()):
                if brand_kana and re.search(r'[a-zA-Z]', brand_kana):
                    # brand_kanaにサブブランド名（G-SHOCK, Baby-G等）が入っている場合
                    # → ブランドフォールバックではなくシリーズエントリとして登録
                    sub_brand_key = brand_kana.upper()
                    if (brand_en, sub_brand_key) not in self.brand_series_map:
                        sub_entry = entry.copy()
                        sub_entry["series_en"] = sub_brand_key
                        sub_entry["brand_kana"] = ""  # カナ検索の汚染を防ぐ
                        self.brand_series_map[(brand_en, sub_brand_key)] = sub_entry
                else:
                    # 通常のブランドフォールバック（brand_kana=カシオ等のカナ）
                    self.brand_fallback_map[brand_en] = entry
            elif series_en:
                self.brand_series_map[(brand_en, series_en)] = entry

            # キーワード登録
            if keywords:
                for kw in keywords.split(","):
                    kw = kw.strip().upper()
                    if kw:
                        self.keyword_map[kw] = (brand_en, series_en)

            # 型番登録（カンマ区切りで複数可）
            if model_numbers:
                for mn in model_numbers.split(","):
                    mn = mn.strip().upper()
                    if mn:
                        self.model_number_map[(brand_en, mn)] = entry

        # === Sheet2: 汎用カテゴリ ===
        ws2 = wb["汎用カテゴリ"]
        header_row = True
        for row in ws2.iter_rows(min_row=1, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]

            if header_row:
                header_row = False
                continue

            # 空行スキップ
            if not vals or all(v == "" for v in vals):
                continue

            # セクションヘッダー行スキップ
            if vals[0].startswith("【"):
                continue

            gender = vals[0] if vals[0] else ""
            movement = vals[1] if len(vals) > 1 else ""
            hand_type = vals[2] if len(vals) > 2 else ""
            category_id = vals[3] if len(vals) > 3 else ""

            if not category_id:
                continue

            self.generic_categories.append({
                "gender": gender,
                "movement": movement,
                "hand_type": hand_type,
                "category_id": category_id,
            })

        wb.close()
        logger.info(
            f"マッピング読込完了: ブランド+シリーズ {len(self.brand_series_map)}件, "
            f"フォールバック {len(self.brand_fallback_map)}件, "
            f"キーワード {len(self.keyword_map)}件, "
            f"型番 {len(self.model_number_map)}件, "
            f"汎用カテゴリ {len(self.generic_categories)}件"
        )

    def lookup(
        self,
        brand_en: str,
        series_en: str = "",
        gender: str = "",
        movement_type: str = "",
        hand_count: str = "",
        model_number: str = "",
    ) -> tuple[str, str, dict | None]:
        """
        カテゴリ番号を検索する。

        Args:
            brand_en: ブランド英字名
            series_en: シリーズ英字名
            gender: 性別（メンズ/レディース/ユニセックス）
            movement_type: ムーブメント種別
            hand_count: 針数
            model_number: 型番（AI解析結果）

        Returns:
            (category_id, match_level, matched_entry)
            match_level: "model_number" / "brand+series" / "brand_only" / "generic" / "unknown"
            matched_entry: 型番マッチ時のエントリ（それ以外はNone）
        """
        brand = brand_en.upper().strip() if brand_en else ""
        series = series_en.upper().strip() if series_en else ""
        model_num = model_number.upper().strip() if model_number else ""

        # === 優先度0: ブランド+型番 完全一致（最優先） ===
        if brand and model_num:
            mn_key = (brand, model_num)
            if mn_key in self.model_number_map:
                entry = self.model_number_map[mn_key]
                if entry["category_id"]:
                    logger.debug(f"型番一致: {brand}+{model_num} → {entry['category_id']}")
                    return entry["category_id"], "model_number", entry

        # === 優先度1: ブランド+シリーズ 完全一致 ===
        if brand and series:
            key = (brand, series)
            if key in self.brand_series_map:
                entry = self.brand_series_map[key]
                if entry["category_id"]:
                    logger.debug(f"完全一致: {brand}+{series} → {entry['category_id']}")
                    return entry["category_id"], "brand+series", None

            # キーワード検索
            if series in self.keyword_map:
                mapped_brand, mapped_series = self.keyword_map[series]
                key2 = (mapped_brand, mapped_series)
                if key2 in self.brand_series_map:
                    entry = self.brand_series_map[key2]
                    if entry["category_id"]:
                        logger.debug(f"キーワード一致: {series} → {mapped_brand}+{mapped_series}")
                        return entry["category_id"], "brand+series", None

            # === 優先度1.5: シリーズ前方一致 ===
            # 例: "G-SHOCK FROGMAN" → "G-SHOCK" にフォールバック
            parts = series.split()
            if len(parts) > 1:
                for i in range(len(parts) - 1, 0, -1):
                    prefix = " ".join(parts[:i])
                    prefix_key = (brand, prefix)
                    if prefix_key in self.brand_series_map:
                        entry = self.brand_series_map[prefix_key]
                        if entry["category_id"]:
                            logger.debug(f"シリーズ前方一致: {brand}+{prefix} → {entry['category_id']}")
                            return entry["category_id"], "brand+series", None

        # === 優先度2: ブランドのみ一致 →「（その他）」 ===
        if brand and brand in self.brand_fallback_map:
            entry = self.brand_fallback_map[brand]
            if entry["category_id"]:
                logger.debug(f"ブランドフォールバック: {brand} → {entry['category_id']}")
                return entry["category_id"], "brand_only", None

        # === 優先度3: 汎用カテゴリ ===
        if gender or movement_type:
            cat_id = self._lookup_generic(gender, movement_type, hand_count)
            if cat_id:
                logger.debug(f"汎用カテゴリ: {gender}/{movement_type}/{hand_count} → {cat_id}")
                return cat_id, "generic", None

        # === 優先度4: 不明 ===
        logger.debug(f"カテゴリ未確定: {brand}/{series}")
        return "", "unknown", None

    def _lookup_generic(self, gender: str, movement: str, hand_count: str) -> str:
        """汎用カテゴリから検索"""
        # 針数の正規化
        hand_type = self._normalize_hand_type(hand_count)

        for cat in self.generic_categories:
            # 性別一致チェック
            if cat["gender"] and gender and cat["gender"] != gender:
                continue
            # ムーブメント一致チェック
            if cat["movement"] and movement and cat["movement"] != movement:
                continue
            # 針タイプ一致チェック
            if cat["hand_type"] and hand_type and cat["hand_type"] != hand_type:
                continue
            # 全条件一致（または条件なし）
            if cat["category_id"]:
                return cat["category_id"]

        # フォールバック: ムーブメント「その他」で再検索
        for cat in self.generic_categories:
            if cat["gender"] and gender and cat["gender"] != gender:
                continue
            if cat["movement"] and movement and cat["movement"] != movement:
                continue
            if cat["hand_type"] == "その他" and cat["category_id"]:
                return cat["category_id"]

        return ""

    def _normalize_hand_type(self, hand_count: str) -> str:
        """針数文字列を汎用カテゴリの形式に変換"""
        if not hand_count:
            return ""

        hc = hand_count.strip()

        # クロノグラフ判定
        if "クロノ" in hc or "chrono" in hc.lower():
            return "クロノグラフ"

        # 2針
        if "2針" in hc or "2本" in hc:
            return "2針（時、分）"

        # 3針
        if "3針" in hc or "3本" in hc:
            return "3針（時、分、秒）"

        # デジタル
        if "デジタル" in hc or "digital" in hc.lower():
            return "デジタル"

        return ""

    def _load_category_names(self):
        """カテゴリ名xlsxを読み込む"""
        category_name_file = CATEGORY_NAME_FILE
        if not category_name_file.exists():
            logger.warning(f"カテゴリ名ファイルが見つかりません: {category_name_file}")
            return

        wb = openpyxl.load_workbook(category_name_file, read_only=True)
        ws = wb.active
        header_row = True
        for row in ws.iter_rows(min_row=1, values_only=True):
            if header_row:
                header_row = False
                continue
            vals = [str(v).strip() if v is not None else "" for v in row]
            cat_id = vals[0] if vals else ""
            cat_name = vals[1] if len(vals) > 1 else ""
            if cat_id and cat_name:
                self.category_name_map[cat_id] = cat_name
        wb.close()
        logger.info(f"カテゴリ名読込完了: {len(self.category_name_map)}件")

    def get_category_name(self, category_id: str) -> str:
        """カテゴリ番号からカテゴリ名を取得"""
        return self.category_name_map.get(category_id.strip(), "")

    def get_brand_kana(self, brand_en: str) -> str:
        """ブランドのカナ表記を取得"""
        brand = brand_en.upper().strip() if brand_en else ""

        # まずブランド+シリーズマップから検索
        for (b, _), entry in self.brand_series_map.items():
            if b == brand and entry["brand_kana"]:
                return entry["brand_kana"]

        # フォールバックマップから検索
        if brand in self.brand_fallback_map:
            return self.brand_fallback_map[brand].get("brand_kana", "")

        return ""

    def get_series_kana(self, brand_en: str, series_en: str) -> str:
        """シリーズのカナ表記を取得"""
        brand = brand_en.upper().strip() if brand_en else ""
        series = series_en.upper().strip() if series_en else ""

        key = (brand, series)
        if key in self.brand_series_map:
            return self.brand_series_map[key].get("series_kana", "")

        return ""
