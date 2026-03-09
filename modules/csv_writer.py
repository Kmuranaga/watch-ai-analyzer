"""
CSV/Excel出力モジュール
テスト版では主要15列を出力する
"""

import csv
import logging
from pathlib import Path
from dataclasses import dataclass

from config import CSV_ENCODING

logger = logging.getLogger(__name__)

# テスト版CSV列定義
COLUMNS = [
    "管理番号",
    "カテゴリ番号",
    "カテゴリ名",
    "タイトル",
    "ブランド英字",
    "ブランドカナ",
    "シリーズ英字",
    "シリーズカナ",
    "型番",
    "素材",
    "防水",
    "ムーブメント",
    "文字盤色",
    "針数",
    "ケース形状",
    "異常内容",
    "処理ステータス",
]


@dataclass
class ProductResult:
    """1商品の処理結果"""
    management_number: str = ""       # 管理番号
    category_id: str = ""             # カテゴリ番号
    category_name: str = ""           # カテゴリ名
    title: str = ""                   # 65文字タイトル
    brand_en: str = ""                # ブランド英字
    brand_kana: str = ""              # ブランドカナ
    series_en: str = ""               # シリーズ英字
    series_kana: str = ""             # シリーズカナ
    model_number: str = ""            # 型番
    material: str = ""                # 素材
    water_resistance: str = ""        # 防水
    movement_type: str = ""           # ムーブメント
    dial_color: str = ""              # 文字盤色
    hand_count: str = ""              # 針数
    case_shape: str = ""              # ケース形状
    abnormality_text: str = ""        # 異常内容
    status: str = "正常"              # 処理ステータス

    def to_row(self) -> list[str]:
        """CSV行として出力"""
        return [
            self.management_number,
            self.category_id,
            self.category_name,
            self.title,
            self.brand_en,
            self.brand_kana,
            self.series_en,
            self.series_kana,
            self.model_number,
            self.material,
            self.water_resistance,
            self.movement_type,
            self.dial_color,
            self.hand_count,
            self.case_shape,
            self.abnormality_text,
            self.status,
        ]


def write_csv(results: list[ProductResult], output_path: Path) -> None:
    """
    処理結果をCSVファイルに出力する。

    Args:
        results: 商品処理結果のリスト
        output_path: 出力先CSVファイルパス
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding=CSV_ENCODING) as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)

        for result in results:
            writer.writerow(result.to_row())

    logger.info(f"CSV出力完了: {output_path} ({len(results)}件)")


def write_excel(results: list[ProductResult], output_path: Path) -> None:
    """
    処理結果をExcelファイルに出力する。

    Args:
        results: 商品処理結果のリスト
        output_path: 出力先Excelファイルパス
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl 未インストール。CSV出力にフォールバックします。")
        csv_path = output_path.with_suffix(".csv")
        write_csv(results, csv_path)
        return

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "解析結果"

    # ヘッダースタイル
    header_font = Font(name="Meiryo", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ヘッダー行
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # データ行
    data_font = Font(name="Meiryo", size=10)
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row_idx, result in enumerate(results, 2):
        row_data = result.to_row()
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # ステータスに応じた色付け
            if result.status != "正常":
                if "エラー" in result.status:
                    cell.fill = error_fill
                else:
                    cell.fill = warning_fill

    # 列幅の自動調整
    column_widths = [14, 14, 50, 50, 16, 14, 18, 14, 16, 12, 10, 14, 10, 8, 14, 30, 20]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else ""].width = width

    # フィルター設定
    ws.auto_filter.ref = ws.dimensions

    # 先頭行を固定
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Excel出力完了: {output_path} ({len(results)}件)")
